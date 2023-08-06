'''
OBJECTIVE	:	copy sheets from industry to the company folder which are excluded.

INPUT VARIABLES	:	filepath	-	directory which will contain two folders - Comapny and Industry.
					Company folder will contain workbook of companies
					Industry folder will have workbook containing industry data
				
OUTPUT  		:	Excluded files copied into the Company file

ASSUMPTION		:	None as such

REQUEST			:	Kindly check the input variables before running the script.

LAST MODIFIED	:	February 19, 2019
'''
from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
import pandas as pnd
import os
import re
import xlrd
import glob

class copytask:

	def __init__(self):
		self.filepath = "DIR"
		temp_list = glob.glob(os.path.join(self.filepath, "Companies", "*.xlsm"))
		temp_list.extend(glob.glob(os.path.join(self.filepath, "Companies", "*.xlsx")))
		self.company_list = temp_list

		temp_list = glob.glob(os.path.join(self.filepath, "Industry", "*.xlsm"))
		temp_list.extend(glob.glob(os.path.join(self.filepath, "Industry", "*.xlsx")))
		self.industry_list = temp_list

		print(self.company_list, self.industry_list, sep="\n")

		self.listofsegments = list()
		self.companyWriter = None 

	def copy_from_industry_to_company(self, sheet_name, workbook_object):
		
		from_df =pnd.read_excel(workbook_object, sheet_name+'_Customers', header=None, engine='xlrd')
		print (sheet_name+'_Customers', ' created')
		from_df.to_excel(self.companyWriter, sheet_name=sheet_name+'_Customers', index=False, header=False, engine='io.excel.xls.writer')
		

		from_df =pnd.read_excel(workbook_object, sheet_name+'_Suppliers', header=None, engine='xlrd')
		print (sheet_name+'_Suppliers', ' created')
		from_df.to_excel(self.companyWriter, sheet_name=sheet_name+'_Suppliers', index=False, header=False, engine='io.excel.xls.writer')

		print ("done")

		#deswb.save(despath)
	

	def check(self):
		
		#Generate a list of industry segments available
		for segment in self.industry_list:
			self.des = load_workbook(segment, keep_vba=True)
			lsd = self.des.sheetnames

			temp_lsd = list()
			#print (lsd)
			for each in lsd:
				ws = self.des[each]
				ws.title = re.sub(' +', ' ', each)
				stringtoadd = ""
				if (len(each)>10):  #to filter invalid names
					name = each[:-10]  #taking only the company name
					stringtoadd = re.sub(' +', ' ', name) #removing extra spaces
				temp_lsd.append(stringtoadd)
			new_lsd = list(set(temp_lsd)) #Keeping only unique values
			self.listofsegments.append(new_lsd)
			self.des.save(segment)
			
		
		cap = len(self.listofsegments)
		
		
		#check sheet in each company file
		for company in self.company_list:
			source = load_workbook(company, read_only=False)
			sh = source.active
			ls = list()
		
			for i in range(1, sh.max_row+1):
				ls.append(re.sub(' +', ' ', sh.cell(i, 1).value))

			excluded = list()
			included = list()

			outputworkbook = Workbook()
			sheet1 = outputworkbook.active

			for sheet in ls:
				count = 0
				segment_index = None
			
				for index, seg in enumerate(self.listofsegments):
					if sheet in seg:
						segment_index = index
						break
					else:
						count+=1
				
				if (count==cap):
					excluded.append(sheet)
				else:
					included.append((sheet, segment_index))

			print ('Included: ', len(included), 'Excluded: ', len(excluded))
		
			for ind, sheet in enumerate(excluded):
				sheet1.cell(ind+1, 1).value = sheet
			
			outputworkbook.save(company)
			#save excluded sheets into the outputworkbook

			self.companyWriter = pnd.ExcelWriter(company, engine='openpyxl') 
			self.companyWriter.book = source
			self.companyWriter.sheets = dict((ws.title, ws) for ws in source.worksheets)

			list_of_workbooks = list()

			for industry in self.industry_list:
				Wb = xlrd.open_workbook(industry)
				list_of_workbooks.append(Wb)

			#print (list_of_workbooks)
		#copy the included sheets
			for pair in included:
				self.copy_from_industry_to_company(pair[0], list_of_workbooks[pair[1]])
			self.companyWriter.save()


if __name__ == '__main__':
	print ("-----STARTED----")
	starttime = datetime.now()
	Obj1 = copytask()
	Obj1.check()
	endtime = datetime.now()
	print ("Time elapsed: ", (endtime-starttime))
	print ("-----DONE----")
	
