from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from copy import copy


class Point:
    def __init__(self, x, y):
        self.x: int = x
        self.y: int = y


def as_text(value):
    return '' if value is None else str(value)


def ws_set_columns_width(ws, float_round_len=3):
    for column_cells in ws.columns:
        max_length = 0

        for cell in column_cells:
            value = cell.value
            if type(cell.value) == float:
                value = round(value, float_round_len)

            cell_len = len(as_text(value))+1

            if type(cell.value) in (int, float):
                cell_len += 2

            if cell_len > max_length:
                max_length = cell_len + 1

        column_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[column_letter].width = min(80, max_length)


def ws_set_number_format(ws, float_round_len=3):
    for i, col in enumerate(ws):
        for cell in col[1:]:
            if type(cell.value) == float:
                cell.number_format = '##0.%s' % ('0' * float_round_len)


def ws_set_header_style(ws, df):
    font = Font(color='00000000', bold=True)
    fill = PatternFill('solid', fgColor='FFFF00')

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    alignment = Alignment(horizontal='center')

    for i, column_name in enumerate(df.columns):
        cell = ws.cell(1, i+1)
        cell.value = column_name
        cell.font = font
        cell.fill = fill
        cell.border = border
        cell.alignment = alignment


def ws_copy_area_to(ws, to: tuple, fro: tuple = (1, 1),
                    width: int = None,
                    height: int = None):
    attrs = ('_style', 'style', 'value', 'font', 'fill', 'border', 'alignment',
             'protection', 'number_format')

    to_ws = to[2] if len(to) > 2 else ws
    fro = Point(*fro)
    to = Point(*to[:2])

    width = ws.max_column if not width else width
    height = ws.max_row if not height else height

    for cx in range(fro.x, width+1):
        for cy in range(fro.y, height+1):
            cell = ws.cell(cy, cx)
            new_cell = to_ws.cell(to.y-1+cy, to.x-1+cx)
            for attr in attrs:
                setattr(new_cell, attr, copy(getattr(cell, attr)))


def excel_merge_vertical(file_names, out_file_name, titles=None):
    wb = Workbook()
    ws = wb.active

    for i, file_name in enumerate(file_names):
        tmp_ws = load_workbook(filename=file_name).active

        y = ws.max_row+2 if ws.max_row > 1 else 1

        title = titles[i] if titles else file_name
        title_cell = ws.cell(y, 1)
        title_cell.value = title
        title_cell.font = Font(color='00000000', bold=True)
        ws_copy_area_to(tmp_ws, to=(1, y+1, ws))

    ws_set_columns_width(ws)
    wb.save(out_file_name)


def excel_merge_sheets(file_names, out_file_name, titles=None):
    wb = Workbook()
    del wb[wb.active.title]

    for i, file_name in enumerate(file_names):
        title = str(titles[i]) if titles else file_name
        ws = wb.create_sheet(title=title)
        tmp_ws = load_workbook(filename=file_name).active
        ws_copy_area_to(tmp_ws, to=(1, 1, ws))
        ws_set_columns_width(ws)

    wb.save(out_file_name)
