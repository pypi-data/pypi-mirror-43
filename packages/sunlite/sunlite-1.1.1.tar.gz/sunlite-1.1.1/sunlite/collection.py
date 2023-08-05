from openpyxl import load_workbook  , Workbook
from sunlite.db import connect
import  json
class xl_connect(object):
       def __init__(self , wb_name = 'new' , overlap = True):
              self.wb = load_workbook(filename=wb_name , read_only=True) if wb_name!='new' else Workbook()
              self.ws = self.wb.active
              self.rows = []
              self.out = {}
              with open('dump', 'r') as f:
                     d = json.loads(f.read())
                     f.close()
              self.overlap = overlap
              self.db = connect(d['DBname'])

       def get_all_rows(self):
              for row in self.ws.rows:
                     self.rows+= [  [x.value for x in row if len(row)!=0]  ]
              return [ x for x in self.rows if x!=[]]

       def get_all_columns(self):
              X = self.get_all_rows()
              result = [[X[j][i] for j in range(len(X))] for i in range(len(X[0]))]
              return  result

       def get_important_row(self , imp='first'):
              if imp=='first' :
                     for x  in self.get_all_rows()[1:]:
                            if len(x)==0:
                                   continue
                            else:
                                   if self.overlap:
                                          pout = []
                                          for i in x[1:]:
                                                 pout += [[self.get_all_rows()[0][1:][ x[1:].index(i)]   , i ] ]
                                          self.out[x[0]] = pout
                                   else:
                                          self.out[x[0]] = x[1:]
              return self.out

       def get_important_cols(self , imp='first'):
              if imp == 'first':
                     for x in self.get_all_columns()[1:]:
                            if len(x) == 0:
                                   continue
                            else:
                                   if self.overlap:
                                          pout = []
                                          for i in x[1:]:
                                                 pout += [[self.get_all_columns()[0][1:][x[1:].index(i)], i]]
                                          self.out[x[0]] = pout
                                   else:
                                          self.out[x[0]] = x[1:]
              return self.out

       def write_to_db(self , type='row') :
              d = self.get_important_row()
              if self.overlap:
                     for item in d:
                            self.db.new(item)
                            for item in [x for x  in d[item] if x!=None]:
                                   self.db.push(item[0] , item[1])
                     d = self.get_important_cols()
                     for item in d:
                            self.db.new(str(item))
                            for item in [x for x in d[item] if x != None]:
                                   self.db.push(item[0], item[1])
              else:
                     d = self.get_important_row()
                     for item in d:
                            self.db.new(str(item))
                            self.db.push(item , d[item])
